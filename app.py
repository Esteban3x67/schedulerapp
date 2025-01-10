from flask import Flask, render_template, jsonify, request
from flask_cors import CORS
from scheduler_core import SchedulerCore
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
from datetime import datetime
from flask import send_file
from openpyxl import load_workbook
import re
import unicodedata
import os
import logging
from logging.handlers import RotatingFileHandler
import sys
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
import json

# Create necessary directories if they don't exist
if not os.path.exists('autosave'):
    os.makedirs('autosave')
if not os.path.exists('logs'):
    os.makedirs('logs')

print("Step 3: About to configure logging")
try:
    # Configure logging but allow Flask output to console
    logging.basicConfig(
        handlers=[
            RotatingFileHandler(
                'logs/app.log', 
                maxBytes=100000, 
                backupCount=3
            ),
            logging.StreamHandler(sys.stdout)
        ],
        level=logging.INFO,
        format='[%(asctime)s] %(levelname)s in %(module)s: %(message)s'
    )
    print("Step 4: Logging configured successfully")
except Exception as e:
    print("Step 4 ERROR: Logging configuration failed")
    print(f"Error details: {str(e)}")
    import traceback
    print(traceback.format_exc())

print("Step 5: About to create Flask app")
app = Flask(
    __name__,
    template_folder='frontend/templates',
    static_folder='frontend/static',
    static_url_path='/static'
)
print("Step 6: Flask app created")

CORS(app, resources={
    r"/api/*": {
        "origins": ["https://scheduler-app-vx70.onrender.com", "http://localhost:5000"],
        "methods": ["GET", "POST", "OPTIONS"]
    }
})

schedulers = {
    'sala': SchedulerCore(),
    'cocina': SchedulerCore(),
    'coperia': SchedulerCore()
}
# Add these constants
GROUP_NAMES = {
    'sala': 'Personal de Sala',
    'cocina': 'Cocina',
    'coperia': 'Copería'
}
# Store hashed password - CHANGE THIS PASSWORD!
MANAGER_PASSWORD_HASH = generate_password_hash('vip123')

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': 'No authorization provided'}), 401
        password = auth_header.split('Basic ')[-1]
        if not check_password_hash(MANAGER_PASSWORD_HASH, password):
            return jsonify({'error': 'Invalid password'}), 401
        return f(*args, **kwargs)
    return decorated

def save_staff_config():
    """Save staff configuration to file"""
    config = {}
    for group in schedulers:
        config[group] = schedulers[group].staff_groups[group]
    
    with open('staff_config.json', 'w') as f:
        json.dump(config, f, indent=2)

def load_staff_config():
    """Load staff configuration from file"""
    try:
        with open('staff_config.json', 'r') as f:
            config = json.load(f)
            for group in schedulers:
                if group in config:
                    schedulers[group].staff_groups[group] = config[group]
    except FileNotFoundError:
        pass  # Use default configuration if file doesn't exist

# Add new route for worker management
@app.route('/api/workers/manage', methods=['POST'])
@requires_auth
def manage_workers():
    """Add, remove, or modify workers"""
    try:
        data = request.get_json()
        action = data.get('action')
        group = data.get('group')
        worker = data.get('worker')
        is_full_time = data.get('is_full_time', True)
        
        if group not in schedulers:
            return jsonify({'success': False, 'error': 'Invalid group'})
            
        scheduler = schedulers[group]
        success = False
        
        if action == 'add':
            success = scheduler.add_worker(worker, is_full_time)
            if not success:
                return jsonify({
                    'success': False, 
                    'error': 'Worker already exists or invalid input'
                })
        elif action == 'remove':
            success = scheduler.remove_worker(worker)
            if not success:
                return jsonify({
                    'success': False, 
                    'error': 'Worker not found'
                })
        
        # Return updated worker list along with success status
        return jsonify({
            'success': True,
            'workers': scheduler.selected_workers
        })
        
    except Exception as e:
        return jsonify({
            'success': False, 
            'error': str(e)
        })
    
@app.route('/')
def index():
    """Show the main page"""
    return render_template('index.html')

@app.route('/api/workers', methods=['GET'])
def get_workers():
    """Return list of workers"""
    group = request.args.get('group', 'sala')
    if group not in schedulers:
        return jsonify({
            'success': False,
            'error': f'Invalid group: {group}'
        }), 400
        
    return jsonify({
        'success': True,
        'workers': schedulers[group].selected_workers
    })
@app.route('/api/generate', methods=['POST'])
def generate():
    """Generate a schedule"""
    try:
        logging.info("Generate endpoint called")
        data = request.get_json()
        logging.info(f"Request data: {data}")
        year = int(data.get('year', 2024))
        month = int(data.get('month', 1))
        group = data.get('group', 'sala')
        
        if group not in schedulers:
            return jsonify({
                'success': False,
                'error': f'Invalid group: {group}'
            }), 400

        scheduler = schedulers[group]
        
        # Generate schedule
        scheduler.initialize_month(year, month)
        scheduler.set_current_group(group)  # Set the correct worker group
        scheduler.assign_night_shifts()
        scheduler.assign_free_sundays()
        scheduler.assign_l_days()
        scheduler.assign_dayshifts()

        # Get the schedule data
        schedule = scheduler.get_month_schedule()
        
        # Add month information
        month_data = {
            'year': year,
            'month': month,
            'days_in_month': calendar.monthrange(year, month)[1],
            'preview_days': 7
        }
        
        return jsonify({
            'success': True,
            'schedule': schedule,
            'month_data': month_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/api/update-shift', methods=['POST'])
def update_shift():
    """Update a single shift"""
    try:
        data = request.get_json()
        worker = data['worker']
        day = int(data['day'])
        shift = data['shift']
        group = data.get('group', 'sala')
        
        if group not in schedulers:
            return jsonify({
                'success': False,
                'error': f'Invalid group: {group}'
            }), 400

        scheduler = schedulers[group]
        worker_index = scheduler.selected_workers.index(worker)

        # Check for T->M violation
        if day > 1:
            prev_shift = scheduler.get_shift(worker_index, day - 1)
            if prev_shift == 'T' and shift == 'M':
                return jsonify({
                    'success': False,
                    'error': 'Cannot assign Morning shift after Tarde shift'
                }), 400

        # Add validation for removing DLs as before
        if shift != 'DL' and scheduler.get_shift(worker_index, day) == 'DL':
            dl_count = sum(1 for d in range(1, scheduler.days_in_month + 1)
                         if d != day and scheduler.get_shift(worker_index, d) == 'DL')
            if dl_count < 2:
                return jsonify({
                    'success': False,
                    'error': 'Cannot remove DL - worker must have 2 DLs per month'
                }), 400

        # Apply the shift
        scheduler.assign_shift(day, worker, shift)
        
        # If this causes any rule violations, let the backend fix them
        if not shift:  # If we're deleting a shift
            scheduler.assign_free_sundays()  # Ensure DL requirements are met
            scheduler.assign_l_days()        # Ensure L day rules are met
        
        # Return the updated schedule
        schedule = scheduler.get_month_schedule()
        return jsonify({
            'success': True,
            'schedule': schedule
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/transfer', methods=['POST'])
def transfer():
    """Transfer preview data to next month"""
    try:
        group = request.get_json().get('group', 'sala')
        if group not in schedulers:
            return jsonify({
                'success': False,
                'error': f'Invalid group: {group}'
            }), 400

        scheduler = schedulers[group]
        # Call the transfer function
        scheduler.transfer_to_next_month()

        # Get the new schedule data
        schedule = scheduler.get_month_schedule()
        
        # Add month information for the new month
        month_data = {
            'year': scheduler.year,
            'month': scheduler.month,
            'days_in_month': scheduler.days_in_month,
            'preview_days': scheduler.preview_days
        }
        
        return jsonify({
            'success': True,
            'schedule': schedule,
            'month_data': month_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/complete-generate', methods=['POST'])
def complete_generate():
    """Complete and generate schedule after transfer"""
    try:
        group = request.get_json().get('group', 'sala')
        if group not in schedulers:
            return jsonify({
                'success': False,
                'error': f'Invalid group: {group}'
            }), 400

        scheduler = schedulers[group]
        
        # Store preview week data temporarily
        preview_data = {}
        for worker_index, worker in enumerate(scheduler.selected_workers):
            worker_shifts = {}
            for day in range(1, 8):
                shift = scheduler.get_shift(worker_index, day)
                if shift:
                    worker_shifts[day] = shift
            preview_data[worker] = worker_shifts

        # Clear schedule but keep core data
        current_year = scheduler.year
        current_month = scheduler.month
        current_days = scheduler.days_in_month
        
        # Reinitialize schedule
        scheduler.initialize_month(current_year, current_month)
        
        # Restore preview week
        for worker, shifts in preview_data.items():
            for day, shift in shifts.items():
                scheduler.assign_shift(day, worker, shift)
        
        # Now follow the same sequence as initial generation
        scheduler.assign_night_shifts_after_transfer()
        scheduler.assign_free_sundays()
        scheduler.assign_l_days()
        scheduler.assign_dayshifts()
        
        # Get the updated schedule
        schedule = scheduler.get_month_schedule()
        
        # Add month information
        month_data = {
            'year': scheduler.year,
            'month': scheduler.month,
            'days_in_month': scheduler.days_in_month,
            'preview_days': 7
        }
        
        return jsonify({
            'success': True,
            'schedule': schedule,
            'month_data': month_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
@app.route('/api/verify-schedule', methods=['GET'])
def verify_schedule():
    """Verify schedule integrity"""
    group = request.args.get('group', 'sala')
    
    if group not in schedulers:
        return jsonify({
            'success': False,
            'error': f'Invalid group: {group}'
        }), 400

    scheduler = schedulers[group]
    violations = []
    
    for worker_index, worker in enumerate(scheduler.selected_workers):
        for day in range(2, scheduler.days_in_month + scheduler.preview_days + 1):
            current_shift = scheduler.get_shift(worker_index, day)
            prev_shift = scheduler.get_shift(worker_index, day - 1)
            
            if prev_shift == 'T' and current_shift == 'M':
                violations.append({
                    'worker': worker,
                    'day': day,
                    'error': 'M shift after T shift',
                    'prev_shift': prev_shift,
                    'current_shift': current_shift
                })
    
    return jsonify({
        'success': True,
        'violations': violations,
        'total_violations': len(violations)
    })

@app.route('/api/verify-dl-counts', methods=['GET'])
def verify_dl_counts():
    """Verify DL counts for each worker in the current month"""
    try:
        group = request.args.get('group', 'sala')
        if group not in schedulers:
            return jsonify({
                'success': False,
                'error': f'Invalid group: {group}'
            }), 400

        scheduler = schedulers[group]
        
        # Get current month's Sundays
        sundays = []
        for day in range(1, scheduler.days_in_month + 1):
            if calendar.weekday(scheduler.year, scheduler.month, day) == 6:  # 6 is Sunday
                sundays.append(day)
        
        # Count DLs for each worker
        dl_status = []
        for worker_index, worker in enumerate(scheduler.selected_workers):
            # Skip part-time workers in DL verification
            if scheduler.is_part_time(worker):
                continue
                
            dl_count = 0
            dl_days = []
            
            for sunday in sundays:
                if scheduler.get_shift(worker_index, sunday) == "DL":
                    dl_count += 1
                    dl_days.append(sunday)
            
            dl_status.append({
                'worker': worker,
                'dl_count': dl_count,
                'dl_days': dl_days,
                'needs_more': dl_count < 2
            })
        
        return jsonify({
            'success': True,
            'dl_status': dl_status,
            'month_sundays': sundays
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def get_column_letter(n):
    """Convert number to Excel column letter (1='A', 27='AA', etc.)"""
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    try:
        wb = Workbook()
        
        # Get current month data from sala group (primary group)
        primary_scheduler = schedulers['sala']
        year = primary_scheduler.year
        month = primary_scheduler.month
        month_names = ["January", "February", "March", "April", "May", "June",
                      "July", "August", "September", "October", "November", "December"]
        
        # Create worksheets for each group
        for group_name in ['sala', 'cocina', 'coperia']:
            scheduler = schedulers[group_name]
            if group_name == 'sala':
                ws = wb.active
                ws.title = 'Personal de Sala'
            else:
                ws = wb.create_sheet(title=GROUP_NAMES.get(group_name, group_name))
            
            # Calculate last column
            last_col = get_column_letter(scheduler.days_in_month + 9)
            
            # Set column widths
            for col in range(1, scheduler.days_in_month + 10):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 4
            ws.column_dimensions['A'].width = 12
            
            # Set the title
            ws.merge_cells(f'A1:{last_col}1')
            ws['A1'] = f"{month_names[month-1]} {year}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Set up styles
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            shift_colors = {
                'N': 'D3D3D3',
                'LN': 'D3D3D3',
                '10N': 'D3D3D3',
                '10LN': 'D3D3D3',
                'L': 'FFFF99',
                'SL': 'FFFF99',
                'DL': '90EE90',
                'M': 'ADD8E6',
                'M4': 'ADD8E6',
                'T': 'FFA07A',
                '2T': 'FFA07A',
                'I': 'E6E6FA'
            }
            
            days = ['D', 'L', 'M', 'X', 'J', 'V', 'S']
            
            # Write headers
            ws.merge_cells('A2:A3')
            ws['A2'] = 'Workers'
            ws['A2'].font = Font(bold=True)
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A2'].border = thin_border
            
            # Write days
            for day in range(1, scheduler.days_in_month + 1):
                col = get_column_letter(day + 1)
                date = f"{year}-{month:02d}-{day:02d}"
                day_name = days[datetime.strptime(date, '%Y-%m-%d').weekday()]
                ws[f'{col}2'] = day_name
                ws[f'{col}3'] = day
                ws[f'{col}2'].font = Font(bold=True)
                ws[f'{col}3'].font = Font(bold=True)
                ws[f'{col}2'].alignment = Alignment(horizontal='center')
                ws[f'{col}3'].alignment = Alignment(horizontal='center')
                ws[f'{col}2'].border = thin_border
                ws[f'{col}3'].border = thin_border
            
            # Add separator and preview days
            separator_col = get_column_letter(scheduler.days_in_month + 2)
            ws[f'{separator_col}2'] = '║'
            ws[f'{separator_col}3'] = '║'
            ws[f'{separator_col}2'].border = thin_border
            ws[f'{separator_col}3'].border = thin_border
            
            for day in range(1, 8):
                col = get_column_letter(scheduler.days_in_month + 2 + day)
                next_month = month + 1 if month < 12 else 1
                next_year = year + 1 if month == 12 else year
                date = f"{next_year}-{next_month:02d}-{day:02d}"
                day_name = days[datetime.strptime(date, '%Y-%m-%d').weekday()]
                ws[f'{col}2'] = day_name
                ws[f'{col}3'] = day
                ws[f'{col}2'].font = Font(bold=True)
                ws[f'{col}3'].font = Font(bold=True)
                ws[f'{col}2'].alignment = Alignment(horizontal='center')
                ws[f'{col}3'].alignment = Alignment(horizontal='center')
                ws[f'{col}2'].border = thin_border
                ws[f'{col}3'].border = thin_border
            
            # Hours column
            hours_col = get_column_letter(scheduler.days_in_month + 10)
            ws[f'{hours_col}2'] = 'Total'
            ws[f'{hours_col}3'] = 'Hours'
            ws[f'{hours_col}2'].font = Font(bold=True)
            ws[f'{hours_col}3'].font = Font(bold=True)
            ws[f'{hours_col}2'].border = thin_border
            ws[f'{hours_col}3'].border = thin_border
            
            # Write schedule data
            schedule = scheduler.get_month_schedule()
            for idx, worker in enumerate(schedule):
                row = idx + 4
                
                # Worker name
                ws[f'A{row}'] = worker['name']
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = thin_border
                
                # Write shifts for current month
                for day in range(1, scheduler.days_in_month + 1):
                    col = get_column_letter(day + 1)
                    shift = worker['shifts'].get(day, '')
                    ws[f'{col}{row}'] = shift
                    ws[f'{col}{row}'].border = thin_border
                    if shift in shift_colors:
                        ws[f'{col}{row}'].fill = PatternFill(start_color=shift_colors[shift],
                                                            end_color=shift_colors[shift],
                                                            fill_type='solid')
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
                
                # Write separator
                separator_col = get_column_letter(scheduler.days_in_month + 2)
                ws[f'{separator_col}{row}'] = '║'
                ws[f'{separator_col}{row}'].border = thin_border
                
                # Write preview days
                for day in range(1, 8):
                    col = get_column_letter(scheduler.days_in_month + 2 + day)
                    actual_day = scheduler.days_in_month + day
                    shift = worker['shifts'].get(actual_day, '')
                    ws[f'{col}{row}'] = shift
                    ws[f'{col}{row}'].border = thin_border
                    if shift in shift_colors:
                        ws[f'{col}{row}'].fill = PatternFill(start_color=shift_colors[shift],
                                                            end_color=shift_colors[shift],
                                                            fill_type='solid')
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
                
                # Write total hours
                ws[f'{hours_col}{row}'] = worker['total_hours']
                ws[f'{hours_col}{row}'].border = thin_border
                ws[f'{hours_col}{row}'].alignment = Alignment(horizontal='center')
        
        # Save to memory buffer
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'schedule_{month_names[month-1]}_{year}.xlsx'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def process_worker_list(worksheet_name, ws, scheduler):
    """Process workers from Excel and merge with existing configuration"""
    processed_workers = set()
    
    # Start from row 4 (after headers) until we hit counter rows
    for row in range(4, ws.max_row + 1):
        worker_name = ws.cell(row=row, column=1).value
        if not worker_name or worker_name in ['Morning:', 'Afternoon:', 'Night:']:
            break
            
        processed_workers.add(worker_name)
        
        # If worker isn't in either list, add them as full-time
        if (worker_name not in scheduler.staff_groups[scheduler.current_group]['workers_full_time'] and 
            worker_name not in scheduler.staff_groups[scheduler.current_group]['workers_part_time']):
            scheduler.add_worker(worker_name, is_full_time=True)
    
    return processed_workers

@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
            
        file = request.files['file']
        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Invalid file format. Please upload an Excel file'}), 400
            
        # Load workbook
        wb = load_workbook(file)
        
        result_data = {}
        
        # Process each worksheet (group)
        for group_name in ['sala', 'cocina', 'coperia']:
            print(f"\nProcessing group: {group_name}")
            print(f"Available worksheets in Excel: {wb.sheetnames}")
            ws = None
            worksheet_name = GROUP_NAMES[group_name]
            print(f"Looking for worksheet: {worksheet_name}")
            
            try:
                ws = wb[worksheet_name]
                print(f"Successfully found worksheet: {worksheet_name}")
                
                # Extract month and year from title
                title = ws['A1'].value
                month_year = title.split()
                month = month_year[0]
                year = int(month_year[1])
                
                # Convert month name to number
                month_names = ["January", "February", "March", "April", "May", "June",
                              "July", "August", "September", "October", "November", "December"]
                month_num = month_names.index(month) + 1
                
                # Initialize scheduler
                scheduler = schedulers[group_name]
                scheduler.initialize_month(year, month_num)
                scheduler.set_current_group(group_name)
                
                # Process workers and merge with existing configuration
                processed_workers = process_worker_list(worksheet_name, ws, scheduler)
                
                # Save updated configuration
                scheduler.save_staff_config()
                
                # Read schedule data as before
                print(f"Reading schedule for {len(scheduler.selected_workers)} workers")
                for worker_index, worker in enumerate(scheduler.selected_workers):
                    if worker not in processed_workers:
                        continue  # Skip workers not in the Excel file
                        
                    print(f"Processing worker: {worker}")
                    for day in range(1, scheduler.days_in_month + 1):
                        col = day + 1  # +1 because first column is worker names
                        cell_value = ws.cell(row=worker_index + 4, column=col).value
                        if cell_value:
                            scheduler.assign_shift(day, worker, cell_value)
                
                # Store result
                result_data[group_name] = {
                    'schedule': scheduler.get_month_schedule(),
                    'month_data': {
                        'year': year,
                        'month': month_num,
                        'days_in_month': calendar.monthrange(year, month_num)[1],
                        'preview_days': 7
                    }
                }
                
            except Exception as e:
                print(f"Error with worksheet {worksheet_name}: {str(e)}")
                continue
        
        # Return all schedules
        return jsonify({
            'success': True,
            'schedules': result_data,
            'current_group': request.args.get('group', 'sala')
        })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500    
@app.route('/api/load-session', methods=['POST'])
def load_session():

    """Load last saved session"""
    try:
        group = request.get_json().get('group', 'sala')
        if group not in schedulers:
            return jsonify({
                'success': False,
                'error': f'Invalid group: {group}'
            }), 400

        scheduler = schedulers[group]
        scheduler.load_last_session()
        
        # Get the loaded schedule
        schedule = scheduler.get_month_schedule()
        
        # Add month information
        month_data = {
            'year': scheduler.year,
            'month': scheduler.month,
            'days_in_month': scheduler.days_in_month,
            'preview_days': scheduler.preview_days
        }
        
        return jsonify({
            'success': True,
            'schedule': schedule,
            'month_data': month_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)